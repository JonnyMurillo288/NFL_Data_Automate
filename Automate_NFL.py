#!/usr/bin/env python
# coding: utf-8
"""
Automate_NFL.py

Clean, modular version of the Jupyter-derived workflow that updates an NFL Excel
workbook with:
- Schedules (2025Schedule sheet) columns B:J
- Depth charts for QB/RB/WR/TE/K ("{POS} Depth" sheets) columns J:O
- Player stats for Pass, Rush, Rec, K, Def sheets

Notes:
- Injury functionality is present but disabled (data source not active for 2025).
- All heavy I/O uses openpyxl to write values without disturbing other cells.
- Name cleaning helpers are consolidated and reused across sheets.

Run this file directly to update all pieces for DEFAULT_YEAR and DEFAULT_OUTPUT_FILE.
"""

from __future__ import annotations

import os
import re
import unicodedata
from datetime import datetime
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import nfl_data_py as nfl

# -----------------------------
# Configuration
# -----------------------------
DEFAULT_OUTPUT_FILE = "NFL_2025.xlsx"
DEFAULT_YEAR = 2025

# Map team abbreviations to full names (used when writing schedule)
abrv_to_full_name = {
    "DAL": "Dallas Cowboys",
    "KC": "Kansas City Chiefs",
    "TB": "Tampa Bay Buccaneers",
    "CIN": "Cincinnati Bengals",
    "MIA": "Miami Dolphins",
    "CAR": "Carolina Panthers",
    "LV": "Las Vegas Raiders",
    "ARI": "Arizona Cardinals",
    "PIT": "Pittsburgh Steelers",
    "NYG": "New York Giants",
    "TEN": "Tennessee Titans",
    "SF": "San Francisco 49ers",
    "DET": "Detroit Lions",
    "HOU": "Houston Texans",
    "BAL": "Baltimore Ravens",
    "MIN": "Minnesota Vikings",
    "WAS": "Washington Commanders",
    "CLE": "Cleveland Browns",
    "JAX": "Jacksonville Jaguars",
    "CHI": "Chicago Bears",
    "NE": "New England Patriots",
    "BUF": "Buffalo Bills",
    "SEA": "Seattle Seahawks",
    "LA": "Los Angeles Rams",
    "DEN": "Denver Broncos",
    "PHI": "Philadelphia Eagles",
    "ATL": "Atlanta Falcons",
    "LAC": "Los Angeles Chargers",
    "GB": "Green Bay Packers",
    "NYJ": "New York Jets",
    "IND": "Indianapolis Colts",
    "NO": "New Orleans Saints",
}

# -----------------------------
# Helpers: names, keys, and writing
# -----------------------------

def clean_name_pretty(s: pd.Series) -> pd.Series:
    """Standardize player names for display and merging.
    - Strip accents, collapse spaces, remove team code suffixes (e.g., " BUF"),
      remove suffixes (JR/SR/II/etc.), and normalize initials.
    """
    def _fix(name: str) -> str:
        if not isinstance(name, str):
            return ""
        name = name.strip()
        name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
        name = re.sub(r"\s+", " ", name)
        name = re.sub(r"\s?[A-Z]{2,4}$", "", name)  # strip glued team code
        name = re.sub(r"\b(JR|SR|II|III|IV|V)\b\.?,?", "", name, flags=re.IGNORECASE)

        def fix_initials(token: str) -> str:
            # Convert "jj" -> "J.J." or keep "A.J." as-is
            if len(token) == 2 and token.isalpha() and token.islower():
                return f"{token[0].upper()}.{token[1].upper()}."
            if len(token) == 2 and token.isalpha() and token.isupper():
                return f"{token[0]}.{token[1]}."
            return token

        name = " ".join(fix_initials(tok) for tok in name.split())
        name = " ".join(tok if "." in tok else tok.capitalize() for tok in name.split())
        return name.strip()

    return s.astype(str).map(_fix)


def name_key(s: pd.Series) -> pd.Series:
    """Create a robust merge key from a name: letters-only, uppercase.
    This ignores spaces, periods, hyphens, apostrophes, and suffixes.
    """
    s = s.astype(str)
    s = s.map(lambda x: unicodedata.normalize("NFKD", x).encode("ascii", "ignore").decode())
    s = s.str.strip()
    s = s.str.replace(r"\s?[A-Z]{2,4}$", "", regex=True)  # trailing team code
    s = s.str.replace(r"\b(JR|SR|II|III|IV|V)\b\.?,?", "", regex=True, flags=re.IGNORECASE)
    s = s.str.replace(r"[^A-Za-z]", "", regex=True).str.upper()
    return s


def write_df_to_sheet(
    df: pd.DataFrame,
    file: str,
    sheet_name: str,
    start_row: int,
    start_col: int,
    only_if_not_na: bool = True,
) -> None:
    """Write a DataFrame to an existing Excel sheet using openpyxl.
    - Writes values cell-by-cell to avoid overwriting unrelated areas.
    - If only_if_not_na, only non-NaN values are written.
    """
    wb = load_workbook(file)
    ws = wb[sheet_name]

    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, value in enumerate(row, start=start_col):
            if (not only_if_not_na) or pd.notna(value):
                ws.cell(row=i, column=j, value=value)

    wb.save(file)


# -----------------------------
# Schedules
# -----------------------------

def _fmt_gametime(value: str) -> str:
    """Format 24-hour HH:MM strings into 12-hour time with AM/PM.
    Returns the original value if formatting fails.
    """
    try:
        if isinstance(value, str) and value.strip():
            return datetime.strptime(value.strip(), "%H:%M").strftime("%I:%M %p")
        return value
    except Exception:
        return value


def update_schedule_sheet(file: str = DEFAULT_OUTPUT_FILE, year: int = DEFAULT_YEAR) -> pd.DataFrame:
    """Update the schedule sheet (columns B:J, rows 2:273 by default) with nflverse data.
    - Reads the existing sheet to preserve shape, then fills mapped columns.
    - Writes back only the target region using openpyxl.
    """
    sheet_name = f"{year}Schedule"
    schedule_sheet = pd.read_excel(file, sheet_name)

    years = [year]
    schedule = nfl.import_schedules(years)

    # Limit our update frame to B:J (index 1:10). Keep number of rows consistent with schedule size.
    n = min(len(schedule_sheet), len(schedule))
    df_update = schedule_sheet.iloc[:n, 1:10].copy()

    # Map into expected columns (existing workbook should already have these columns created).
    df_update["Week"] = schedule["week"].values[:n]
    df_update["Day"] = schedule["weekday"].values[:n]
    df_update["Date"] = schedule["gameday"].values[:n]

    df_update["Team"] = schedule["away_team"].map(abrv_to_full_name).values[:n]
    df_update["Team PTS"] = schedule["away_score"].values[:n]
    df_update["Home or Away"] = "Away"
    df_update["Opposing Team"] = schedule["home_team"].map(abrv_to_full_name).values[:n]
    df_update["OPP PTS"] = schedule["home_score"].values[:n]
    df_update["Time"] = schedule["gametime"].map(_fmt_gametime).values[:n]

    # Write DataFrame values into B2:J(1+n)
    write_df_to_sheet(df_update, file=file, sheet_name=sheet_name, start_row=2, start_col=2, only_if_not_na=False)
    return df_update


# -----------------------------
# Depth charts
# -----------------------------

def get_depth_by_team(depth_charts: pd.DataFrame, pos_abb: str, max_slots: int = 5) -> pd.DataFrame:
    """Build a depth chart table for a given position abbreviation.
    Returns columns: Team, {POS} #1...{POS} #max_slots
    """
    # Sheet label uses K for PK
    sheet_pos = "K" if pos_abb == "PK" else pos_abb
    rows = []
    subset = depth_charts.loc[depth_charts.pos_abb == pos_abb]

    for abv in abrv_to_full_name.keys():
        # Display 'Team' column as the last word of the full team name (e.g., "Cowboys")
        team_display = abrv_to_full_name.get(abv, "").split()[-1]
        df_team = subset.loc[subset.team == abv].sort_values("pos_rank")
        names = df_team.player_name.dropna().astype(str).unique().tolist()
        # Keep max_slots players, pad with empty strings if fewer
        names = (names + [""] * max_slots)[:max_slots]
        rows.append([team_display] + names)

    columns = ["Team"] + [f"{sheet_pos} #{i}" for i in range(1, max_slots + 1)]
    return pd.DataFrame(rows, columns=columns)


def update_depth_charts(file: str = DEFAULT_OUTPUT_FILE, years: List[int] | None = None) -> None:
    """Update depth chart sheets for QB/RB/WR/TE/K.
    - Writes data into columns J:O (start_col=10), starting at row 2.
    - Uses nflverse depth charts with pos_abb in [QB, RB, WR, TE, PK].
    """
    if years is None:
        years = [DEFAULT_YEAR]

    depth = nfl.import_depth_charts(years)
    # Positions in the source; map PK -> K for sheet naming
    positions = ["QB", "RB", "WR", "TE", "PK"]
    for pos in positions:
        sheet_pos = "K" if pos == "PK" else pos
        dp = get_depth_by_team(depth, pos)
        # Write to e.g., "QB Depth", "K Depth"
        write_df_to_sheet(dp.iloc[:, :6], file=file, sheet_name=f"{sheet_pos} Depth", start_row=2, start_col=10)


# -----------------------------
# Player stats import and preparation
# -----------------------------

def import_seasonal_data(years: List[int], s_type: str = "REG") -> pd.DataFrame:
    """Import seasonal player data from nflverse parquet.
    - s_type in (REG, ALL, POST)
    - Returns player-week records; downstream code aggregates or maps as needed.
    """
    if not isinstance(years, (list, range)):
        raise ValueError("years input must be list or range.")
    if min(years) < 1999:
        raise ValueError("Data not available before 1999.")
    if s_type not in ("REG", "ALL", "POST"):
        raise ValueError("Only REG, ALL, POST allowed for s_type.")

    url = r"https://github.com/nflverse/nflverse-data/releases/download/stats_player/stats_player_reg_{0}.parquet"
    data = pd.concat([pd.read_parquet(url.format(x), engine="auto") for x in years])

    if s_type != "ALL":
        data = data[(data["season_type"] == s_type)]

    # Custom receiving-style share features (optional usage downstream)
    season_stats = data.copy()
    season_stats["ppr_pts"] = season_stats["fantasy_points_ppr"]
    season_stats["tgt_sh"] = season_stats["targets"] / season_stats["attempts"]
    season_stats["ay_sh"] = season_stats["receiving_air_yards"] / season_stats["passing_air_yards"]
    season_stats["yac_sh"] = season_stats["receiving_yards_after_catch"] / season_stats["passing_yards_after_catch"]
    season_stats["wopr"] = season_stats["target_share"] * 1.5 + season_stats["air_yards_share"] * 0.8
    season_stats["ry_sh"] = season_stats["receiving_yards"] / season_stats["passing_yards"]
    season_stats["rtd_sh"] = season_stats["receiving_tds"] / season_stats["passing_yards"]
    season_stats["rfd_sh"] = season_stats["receiving_first_downs"] / season_stats["passing_first_downs"]
    season_stats["rtdfd_sh"] = (
        (season_stats["receiving_tds"] + season_stats["receiving_first_downs"]) /
        (season_stats["passing_tds"] + season_stats["passing_first_downs"])
    )
    season_stats["dom"] = (season_stats["ry_sh"] + season_stats["rtd_sh"]) / 2
    season_stats["w8dom"] = season_stats["ry_sh"] * 0.8 + season_stats["rtd_sh"] * 0.2
    season_stats["yptmpa"] = season_stats["receiving_yards"] / season_stats["attempts"]
    season_stats["ppr_sh"] = season_stats["fantasy_points_ppr"] / season_stats["ppr_pts"]

    return season_stats


def prepare_player_datasets(year: int) -> Dict[str, pd.DataFrame]:
    """Prepare merged player-level datasets for different position groups.
    Returns a dict with keys: passing, rushing, receiving, kicking, defense."""
    df = import_seasonal_data([year])
    players = nfl.import_players()

    merged = df.merge(
        players[["gsis_id", "common_first_name", "last_name", "position"]],
        how="left",
        left_on="player_id",
        right_on="gsis_id",
        indicator=False,
    )
    merged["Name"] = merged["common_first_name"].fillna("") + " " + merged["last_name"].fillna("")
    merged = merged.rename(columns={"position_x": "position"})

    datasets = {
        "passing": merged.loc[merged.position == "QB"].copy(),
        "rushing": merged.loc[merged.position == "RB"].copy(),
        "receiving": merged.loc[merged.position.isin(["RB", "WR", "TE"])].copy(),
        "kicking": merged.loc[merged.position == "K"].copy(),
        "defense": merged.loc[merged.position.isin(["DL", "OL", "LB", "DB"])].copy(),
    }
    return datasets


# -----------------------------
# Sheet updaters (Pass, Rush, Rec, K, Def)
# -----------------------------

def update_pass_sheet(datasets: Dict[str, pd.DataFrame], file: str = DEFAULT_OUTPUT_FILE) -> None:
    """Update Pass sheet using mapped columns from passing dataset.
    Writes starting at A2.
    """
    passing = datasets["passing"].copy()
    passing_sheet = pd.read_excel(file, "Pass")

    # Clean display names
    if "Name" not in passing.columns and {"first_name", "last_name"} <= set(passing.columns):
        passing["Name"] = (
            passing["first_name"].fillna("").str.strip() + " " + passing["last_name"].fillna("").str.strip()
        )
    passing["Name"] = clean_name_pretty(passing["Name"])
    passing_sheet["Name"] = clean_name_pretty(passing_sheet["Name"])

    # Map source -> destination
    map_to_sheet = {
        "completions": "COMP",
        "attempts": "ATT",
        "passing_yards": "YDS",
        "passing_tds": "TD",
        "interceptions": "INT",
        "sacks": "SCK",
        "carries": "ATT.1",
        "rushing_yards": "YDS.1",
        "rushing_tds": "TD.1",
        "rushing_fumbles": "FUM",
        "games": "G",
        # Include if present
        "qbr": "QBR",
    }

    src_cols = [c for c in map_to_sheet if c in passing.columns]
    tmp = passing[["Name"] + src_cols].copy()
    for c in src_cols:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce")

    merged = passing_sheet.merge(tmp, on="Name", how="left", suffixes=("", "_src"))
    for src, dest in map_to_sheet.items():
        if src in merged.columns and dest in merged.columns:
            merged[dest] = merged[src].combine_first(merged[dest])
            merged.drop(columns=src, inplace=True)

    write_df_to_sheet(merged, file=file, sheet_name="Pass", start_row=2, start_col=1)


def update_rush_sheet(datasets: Dict[str, pd.DataFrame], file: str = DEFAULT_OUTPUT_FILE) -> None:
    """Update Rush sheet using robust name-key merge.
    Workbook has header offset (header=2). Writes starting at A4.
    """
    rushing = datasets["rushing"].copy()
    rush_sheet = pd.read_excel(file, "Rush", header=2)

    rushing["Name"] = clean_name_pretty(rushing.get("Name", pd.Series(dtype=str)))
    rush_sheet["Name"] = clean_name_pretty(rush_sheet["Name"])

    rushing["__key__"] = name_key(rushing["Name"])
    rush_sheet["__key__"] = name_key(rush_sheet["Name"])

    map_to_sheet = {
        "carries": "ATT",
        "rushing_yards": "YDS",
        "rushing_tds": "TD",
        "games": "G",
        "rushing_fumbles": "TOT",
        "rushing_fumbles_lost": "LOST",
    }

    src_cols = [c for c in map_to_sheet if c in rushing.columns]
    tmp = rushing[["__key__"] + src_cols].copy()
    for c in src_cols:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce")

    # Ensure destination columns exist
    for dest in map_to_sheet.values():
        if dest not in rush_sheet.columns:
            rush_sheet[dest] = np.nan

    merged = rush_sheet.merge(tmp, on="__key__", how="left", suffixes=("", "_src"))
    for src, dest in map_to_sheet.items():
        if src in merged.columns and dest in merged.columns:
            merged[dest] = np.where(merged[src].notna(), merged[src], merged[dest])
            merged.drop(columns=src, inplace=True)

    # Finalize and write (drop helper key)
    merged = merged.drop(columns="__key__", errors="ignore")
    write_df_to_sheet(merged, file=file, sheet_name="Rush", start_row=4, start_col=1)


def update_rec_sheet(datasets: Dict[str, pd.DataFrame], file: str = DEFAULT_OUTPUT_FILE) -> None:
    """Update Rec sheet. Workbook header offset (header=1). Writes starting at A3."""
    receiving = datasets["receiving"].copy()
    rec_sheet = pd.read_excel(file, "Rec", header=1)

    # Build Name if not present
    if "Name" not in receiving.columns:
        if {"first_name", "last_name"} <= set(receiving.columns):
            receiving["Name"] = (
                receiving["first_name"].fillna("").str.strip() + " " + receiving["last_name"].fillna("").str.strip()
            )
        elif {"common_first_name", "last_name"} <= set(receiving.columns):
            receiving["Name"] = (
                receiving["common_first_name"].fillna("").str.strip()
                + " "
                + receiving["last_name"].fillna("").str.strip()
            )

    # Clean
    receiving["Name"] = clean_name_pretty(receiving["Name"])
    rec_sheet["Name"] = clean_name_pretty(rec_sheet["Name"])

    map_to_sheet = {
        "receptions": "REC",
        "targets": "TAR",
        "receiving_yards": "YDS",
        "receiving_tds": "TD",
        "games": "G",
        "receiving_fumbles": "TOT",
        "receiving_fumbles_lost": "LOST",
    }

    src_cols = [c for c in map_to_sheet if c in receiving.columns]
    tmp = receiving[["Name"] + src_cols].copy()
    for c in src_cols:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce")

    merged = rec_sheet.merge(tmp, on="Name", how="left", suffixes=("", "_src"))
    for src, dest in map_to_sheet.items():
        if src in merged.columns and dest in merged.columns:
            merged[dest] = merged[src].combine_first(merged[dest])
            merged.drop(columns=src, inplace=True)

    write_df_to_sheet(merged, file=file, sheet_name="Rec", start_row=3, start_col=1)


def update_k_sheet(datasets: Dict[str, pd.DataFrame], file: str = DEFAULT_OUTPUT_FILE) -> None:
    """Update K sheet. Workbook header offset (header=2). Writes starting at A4."""
    kicking = datasets["kicking"].copy()
    k_sheet = pd.read_excel(file, "K", header=2)

    # Build Name if needed
    if "Name" not in kicking.columns:
        if {"first_name", "last_name"} <= set(kicking.columns):
            kicking["Name"] = kicking["first_name"].fillna("").str.strip() + " " + kicking["last_name"].fillna("").str.strip()
        elif {"common_first_name", "last_name"} <= set(kicking.columns):
            kicking["Name"] = (
                kicking["common_first_name"].fillna("").str.strip() + " " + kicking["last_name"].fillna("").str.strip()
            )

    kicking["Name"] = clean_name_pretty(kicking["Name"])
    k_sheet["Name"] = clean_name_pretty(k_sheet["Name"])

    map_to_sheet = {
        "games": "G",
        "xp_made": "XPM",
        "xp_attempts": "XPA",
        "fg_made": "FGM",
        "fg_attempts": "FGA",
        "fg_made_1_19": "FGM.1",
        "fg_attempts_1_19": "FGA.1",
        "fg_made_20_29": "FGM.2",
        "fg_attempts_20_29": "FGA.2",
        "fg_made_30_39": "FGM.3",
        "fg_attempts_30_39": "FGA.3",
        "fg_made_40_49": "FGM.4",
        "fg_attempts_40_49": "FGA.4",
    }

    src_cols = [c for c in map_to_sheet if c in kicking.columns]
    tmp = kicking[["Name"] + src_cols].copy()
    for c in src_cols:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce")

    merged = k_sheet.merge(tmp, on="Name", how="left", suffixes=("", "_src"))
    for src, dest in map_to_sheet.items():
        if src in merged.columns and dest in merged.columns:
            merged[dest] = merged[src].combine_first(merged[dest])
            merged.drop(columns=src, inplace=True)

    write_df_to_sheet(merged, file=file, sheet_name="K", start_row=4, start_col=1)


def update_def_sheet(datasets: Dict[str, pd.DataFrame], file: str = DEFAULT_OUTPUT_FILE) -> None:
    """Update Def sheet. Workbook header offset (header=1). Writes starting at A3."""
    defense = datasets["defense"].copy()
    def_sheet = pd.read_excel(file, "Def", header=1)

    # Build Name if needed
    if "Name" not in defense.columns:
        if {"first_name", "last_name"} <= set(defense.columns):
            defense["Name"] = defense["first_name"].fillna("").str.strip() + " " + defense["last_name"].fillna("").str.strip()
        elif {"common_first_name", "last_name"} <= set(defense.columns):
            defense["Name"] = (
                defense["common_first_name"].fillna("").str.strip()
                + " "
                + defense["last_name"].fillna("").str.strip()
            )

    defense["Name"] = clean_name_pretty(defense["Name"])  # correct: clean defense names
    def_sheet["Name"] = clean_name_pretty(def_sheet["Name"])  # also clean sheet names

    map_to_sheet = {
        "def_tackles_solo": "SOLO",
        "def_tackle_assists": "AST",
        "def_tackles_with_assist": "TOT TCKL",
        "def_sacks": "SACK",
        "def_fumbles": "FUM Forced",
        "fumble_recovery_own": "FUM Rec",
        "def_interceptions": "INT",
        "def_pass_defended": "PASS DEF",
    }

    src_cols = [c for c in map_to_sheet if c in defense.columns]
    tmp = defense[["Name"] + src_cols].copy()
    for c in src_cols:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce")

    merged = def_sheet.merge(tmp, on="Name", how="left", suffixes=("", "_src"))
    for src, dest in map_to_sheet.items():
        if src in merged.columns and dest in merged.columns:
            merged[dest] = merged[src].combine_first(merged[dest])
            merged.drop(columns=src, inplace=True)

    write_df_to_sheet(merged, file=file, sheet_name="Def", start_row=3, start_col=1)


# -----------------------------
# Injury (disabled)
# -----------------------------

def update_injury_report(years: List[int] | None = None, file: str = DEFAULT_OUTPUT_FILE) -> None:
    """Disabled placeholder for injury report.
    2025 injury data source is not currently available.
    """
    raise RuntimeError(
        "Injury update is disabled. Activate once injuries_{year}.parquet is available."
    )


# -----------------------------
# Orchestration
# -----------------------------

def run_all(output_file: str = DEFAULT_OUTPUT_FILE, year: int = DEFAULT_YEAR) -> None:
    """Run the full update pipeline (excluding injuries)."""
    print(f"Updating schedule for {year} -> {output_file}")
    update_schedule_sheet(file=output_file, year=year)

    print("Updating depth charts")
    update_depth_charts(file=output_file, years=[year])

    print("Preparing player datasets")
    datasets = prepare_player_datasets(year)

    print("Updating Pass sheet")
    update_pass_sheet(datasets, file=output_file)

    print("Updating Rush sheet")
    update_rush_sheet(datasets, file=output_file)

    print("Updating Rec sheet")
    update_rec_sheet(datasets, file=output_file)

    print("Updating K sheet")
    update_k_sheet(datasets, file=output_file)

    print("Updating Def sheet")
    update_def_sheet(datasets, file=output_file)

    print("All updates complete.")


if __name__ == "__main__":
    # Execute the pipeline with defaults
    run_all()
